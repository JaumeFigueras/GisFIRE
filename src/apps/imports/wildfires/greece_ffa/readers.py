#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Reads the Greek Fire Service workbooks into one record per fire.

:func:`read_workbook` yields a :class:`Sheet` per year found in a ``.xlsx``, each
carrying its :class:`FireRecord` iterator. The 2000-2012 file holds thirteen
sheets and every other file holds one, so a *file* is not the unit of anything
here — a **year** is, which is also the unit the importer replaces.

Why the columns are found and not counted
------------------------------------------

The sheets have 16, 17, 31, 32, 36, 38 or 39 columns depending on the year, in six
different arrangements (see :mod:`src.providers.greece_ffa`), so every field is
located by matching its header rather than by position. The match runs through
:func:`~src.providers.greece_ffa.normalise_column`, which is what makes
``ΒΥΤΙΟ- ΦΟΡΑ`` and ``ΒΥΤΙΟΦΟΡΑ`` one column, and what makes the Latin-``A``
``A/A ENGAGE`` of the 2025 file the same column as the Greek-``Α`` one of every
other year.

:data:`COLUMNS` is the whole mapping, and a header this reader does not know is
reported rather than ignored — see :attr:`Sheet.unknown_columns`. A column that
quietly disappeared from a future publication would otherwise import as twenty-six
years of nulls.

Where the header is
-------------------

Not at a fixed row. Every sheet has a banner row above the names
(``ΚΑΜΜΕΝΗ ΕΚΤΑΣΗ``, ``ΠΡΟΣΩΠΙΚΟ``, …) and the 2025 file has a title and a year
above that. :func:`find_header` looks for the first row that names at least
:data:`MIN_KNOWN_COLUMNS` columns this reader knows, which is a property of the
row rather than of the file and so needs no per-year table.

Which year a sheet is
---------------------

The sheet name, when it is four digits — every file but one. The 2025 file calls
its sheet ``Sheet0`` and prints the year in a cell above the header
(``Για το ΕΤΟΣ:`` ``2025``), so :func:`find_year` falls back to the first
plausible year it finds in the rows before the header. A sheet whose year cannot
be established is refused outright, because the year is what the import replaces
and getting it wrong would delete the wrong one.

Reading with ``openpyxl``, and with ``data_only``
--------------------------------------------------

Unlike the EGIF reader next door, which parses the stored XML itself, this one
uses ``openpyxl`` — for three things that reader would have to reimplement:
per-sheet access in a thirteen-sheet workbook, multi-row headers, and above all
**cached formula results**. The 2022-2024 coordinate columns are not numbers but
``VLOOKUP`` calls into a helper sheet; ``data_only=True`` returns the values Excel
last stored for them, which is what makes the helper sheets unnecessary.

The workbooks are 1-14 MB, so ``read_only=True`` streaming is a convenience rather
than a necessity, and it is used anyway: it keeps a 260,000-row import to a
constant memory footprint.

Nothing is validated here
-------------------------

A reader's job is to say what the file contains, not whether it is usable. A row
whose start date will not parse arrives with ``start_date`` set to ``None`` and
the raw value in :attr:`FireRecord.problems`, and it is the importer that decides
such a row cannot be stored. Keeping the two apart is what lets the importer
report *which* row was dropped and why.
"""

from __future__ import annotations

import dataclasses
import datetime
import re
import typing

from pathlib import Path

import openpyxl

from src.providers import greece_ffa
from src.providers.greece_ffa import normalise_column

#: Sheets that are not a year of fires. ``engage``/``engagexy`` are the helper
#: tables the 2022-2024 coordinate formulas look into — already resolved in the
#: cached values, so reading them would import every coordinate twice — and
#: ``SQL Statement`` is a one-cell note the 2023 publication left behind.
IGNORED_SHEETS = frozenset({"ENGAGE", "ENGAGEXY", "SQLSTATEMENT"})

#: How many rows to look at before giving up on finding a header.
MAX_HEADER_ROW = 8

#: How many recognised names a row needs before it is taken for the header row.
#:
#: The banner row above it names four groups at most (``ΚΑΜΜΕΝΗ ΕΚΤΑΣΗ``,
#: ``ΠΡΟΣΩΠΙΚΟ``, ``ΟΧΗΜΑΤΑ``, ``ΕΝΑΕΡΙΑ``), none of which is in :data:`COLUMNS`,
#: and the narrowest real header has sixteen columns — so the gap is wide and the
#: threshold is not a tuned number.
MIN_KNOWN_COLUMNS = 8

#: Years a sheet may plausibly claim, for :func:`find_year`.
PLAUSIBLE_YEARS = (1990, 2100)

#: ``dd/mm/yyyy``, how the 2025 file writes a date. Every other year stores a real
#: Excel datetime.
TEXT_DATE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$")

#: ``HH:MM``, how most years write a time. 2000-2006 and 2008-2010 store a real
#: Excel time instead.
TEXT_TIME = re.compile(r"^\s*(\d{1,2}):(\d{2})(?::(\d{2}))?\s*$")

#: Published header (normalised) to the :class:`FireRecord` field it fills.
#:
#: Every spelling a column has ever been published under maps to the same field,
#: which is the whole point of normalising first: ``ΟΧΗΜ. ΟΤΑ`` (2011-2021) and
#: ``ΟΧΗΜ. ΥΠΗΡΕΣΙΑΚΑ`` (2022 on) are one slot renamed, and
#: ``ΠΕΡΙΟΧΗ - ΤΟΠΟΘΕΣΙΑ`` (2000-2011) is what became ``ΠΕΡΙΟΧΗ`` when
#: ``ΔΙΕΥΘΥΝΣΗ`` split off it in 2012.
COLUMNS: dict[str, str] = {
    # -- identity
    normalise_column("Α/Α ΕΓΓΡΑΦΗΣ"): "record_number",
    normalise_column("Α/Α ENGAGE"): "engage_id",
    normalise_column("X-ENGAGE"): "longitude",
    normalise_column("Y-ENGAGE"): "latitude",
    normalise_column("Κατηγορία Συμβάντος"): "incident_category",
    # -- when
    normalise_column("Ημερ/νία Έναρξης"): "start_date",
    normalise_column("Ώρα Έναρξης"): "start_time",
    normalise_column("Ημερ/νία Κατασβεσης"): "end_date",
    normalise_column("Ώρα Κατάσβεσης"): "end_time",
    # -- where
    normalise_column("Υπηρεσία"): "station_name",
    normalise_column("Νομός"): "prefecture_name",
    normalise_column("Δασαρχείο"): "forest_district_name",
    normalise_column("Δήμος"): "municipality_name",
    normalise_column("Περιοχή"): "locality_name",
    normalise_column("Περιοχή - Τοποθεσία"): "locality_name",
    normalise_column("Διεύθυνση"): "address",
    # -- what burnt, in στρέμματα
    normalise_column("Δάση"): "area_forest",
    normalise_column("Δασική Έκταση"): "area_forest_land",
    normalise_column("Άλση"): "area_grove",
    normalise_column("Χορτ/κές Εκτάσεις"): "area_grassland",
    normalise_column("Καλάμια - Βάλτοι"): "area_reeds_marsh",
    normalise_column("Γεωργικές Εκτάσεις"): "area_agricultural",
    normalise_column("Υπολλείματα Καλλιεργειών"): "area_crop_residue",
    normalise_column("Σκουπιδότοποι"): "area_landfill",
    normalise_column("Σκουπι-δότοποι"): "area_landfill",
    # -- who and what was sent
    normalise_column("ΠΥΡΟΣ. ΣΩΜΑ"): "personnel_fire_service",
    normalise_column("ΠΕΖΟΠΟΡΑ ΤΜΗΜΑΤΑ"): "personnel_infantry_units",
    normalise_column("ΕΘΕΛΟΝΤΕΣ"): "personnel_volunteers",
    normalise_column("ΕΘΕΛΟ-ΝΤΕΣ"): "personnel_volunteers",
    normalise_column("ΣΤΡΑΤΟΣ"): "personnel_army",
    normalise_column("ΑΛΛΕΣ ΔΥΝΑΜΕΙΣ"): "personnel_other",
    normalise_column("ΠΥΡΟΣ. ΟΧΗΜ."): "vehicles_fire_service",
    normalise_column("ΟΧΗΜ. ΟΤΑ"): "vehicles_public_service",
    normalise_column("ΟΧΗΜ. ΥΠΗΡΕΣΙΑΚΑ"): "vehicles_public_service",
    normalise_column("ΒΥΤΙΟΦΟΡΑ"): "vehicles_water_tankers",
    normalise_column("ΒΥΤΙΟ- ΦΟΡΑ"): "vehicles_water_tankers",
    normalise_column("ΜΗΧΑΝΗΜΑΤΑ"): "vehicles_machinery",
    normalise_column("ΜΗΧΑΝΗ-ΜΑΤΑ"): "vehicles_machinery",
    normalise_column("ΕΛΙΚΟΠΤΕΡΑ"): "aircraft_helicopters",
    normalise_column("ΕΛΙΚΟ- ΠΤΕΡΑ"): "aircraft_helicopters",
    normalise_column("Α/Φ CL415"): "aircraft_cl415",
    normalise_column("Α/Φ CL215"): "aircraft_cl215",
    normalise_column("Α/Φ PZL"): "aircraft_pzl",
    normalise_column("Α/Φ GRU."): "aircraft_gru",
    normalise_column("ΜΙΣΘ. ΕΛΙΚΟΠΤ."): "aircraft_leased_helicopters",
    normalise_column("ΜΙΣΘ. ΑΕΡΟΣΚ."): "aircraft_leased_planes",
    normalise_column("ΑΛΛΩΝ ΦΟΡΕΩΝ"): "aircraft_other_agencies",
}

#: Fields holding a burnt area in στρέμματα as published. The importer converts
#: them; the reader does not, so a record still says what the file said.
AREA_FIELDS = ("area_forest", "area_forest_land", "area_grove", "area_grassland",
               "area_reeds_marsh", "area_agricultural", "area_crop_residue",
               "area_landfill")

#: Fields holding a count of people, vehicles or aircraft.
COUNT_FIELDS = ("personnel_fire_service", "personnel_infantry_units",
                "personnel_volunteers", "personnel_army", "personnel_other",
                "vehicles_fire_service", "vehicles_public_service",
                "vehicles_water_tankers", "vehicles_machinery",
                "aircraft_helicopters", "aircraft_cl415", "aircraft_cl215",
                "aircraft_pzl", "aircraft_gru", "aircraft_leased_helicopters",
                "aircraft_leased_planes", "aircraft_other_agencies")


@dataclasses.dataclass(slots=True)
class FireRecord:
    """One published row, with the fields its own year publishes filled in.

    Every field is optional. A year that does not publish a column leaves it
    ``None``, and the importer stores that ``None`` — which is the difference
    between *not published* and *zero*, and the reason no field defaults to 0.

    Attributes
    ----------
    row : int
        The 1-based worksheet row the record came from, for the importer to name
        in a message. This dataset has no identifier for 201,948 of its rows, so
        the row number is frequently the only way to say *which* one was refused.
    problems : list of str
        What could not be read, in words. Empty for a clean record.
    """

    row: int
    problems: list[str] = dataclasses.field(default_factory=list)

    # -- identity
    record_number: int | None = None
    engage_id: int | None = None
    incident_category: str | None = None

    # -- when, as published: a date and a wall-clock time, separately
    start_date: datetime.date | None = None
    start_time: datetime.time | None = None
    end_date: datetime.date | None = None
    end_time: datetime.time | None = None

    # -- where
    longitude: float | None = None
    latitude: float | None = None
    station_name: str | None = None
    prefecture_name: str | None = None
    forest_district_name: str | None = None
    municipality_name: str | None = None
    locality_name: str | None = None
    address: str | None = None

    # -- what burnt, in στρέμματα as published
    area_forest: float | None = None
    area_forest_land: float | None = None
    area_grove: float | None = None
    area_grassland: float | None = None
    area_reeds_marsh: float | None = None
    area_agricultural: float | None = None
    area_crop_residue: float | None = None
    area_landfill: float | None = None

    # -- what was sent
    personnel_fire_service: int | None = None
    personnel_infantry_units: int | None = None
    personnel_volunteers: int | None = None
    personnel_army: int | None = None
    personnel_other: int | None = None
    vehicles_fire_service: int | None = None
    vehicles_public_service: int | None = None
    vehicles_water_tankers: int | None = None
    vehicles_machinery: int | None = None
    aircraft_helicopters: int | None = None
    aircraft_cl415: int | None = None
    aircraft_cl215: int | None = None
    aircraft_pzl: int | None = None
    aircraft_gru: int | None = None
    aircraft_leased_helicopters: int | None = None
    aircraft_leased_planes: int | None = None
    aircraft_other_agencies: int | None = None

    @property
    def start_datetime(self) -> datetime.datetime | None:
        """The published start as one naive local reading, or ``None``.

        Every row in the archive publishes both a start date and a start time, so
        the ``None`` branch is for a file that has stopped doing so rather than for
        anything in the twenty-six years read so far.
        """
        if self.start_date is None:
            return None
        return datetime.datetime.combine(self.start_date,
                                         self.start_time or datetime.time.min)

    @property
    def end_datetime(self) -> datetime.datetime | None:
        """The published extinction as one naive local reading, or ``None``.

        ``None`` when no end **date** was published — 27,183 rows, one in ten —
        including the 586 that publish an extinction *time* and no date, which is
        not enough to name an instant.

        A date with no time (641 rows) is read as local midnight, on the rule the
        rest of the project follows for a provider that publishes a bare date; see
        :mod:`src.data_model.wildfire`.
        """
        if self.end_date is None:
            return None
        return datetime.datetime.combine(self.end_date,
                                         self.end_time or datetime.time.min)

    @property
    def located(self) -> bool:
        """Whether this row publishes a usable coordinate."""
        return greece_ffa.is_located(self.longitude, self.latitude)


@dataclasses.dataclass(slots=True)
class Sheet:
    """One year of fires, as found in a workbook.

    Attributes
    ----------
    year : int
        The year the sheet is filed under, from its name or from the cell above
        its header. This is what the import replaces.
    name : str
        The worksheet name, stored as
        :attr:`~src.providers.greece_ffa.wildfire.GreeceFfaWildfire.source_sheet`.
    header_row : int
        The 1-based row the column names were found on.
    rows : int
        How many data rows the sheet has, counted from the worksheet dimensions so
        that a progress bar can show a percentage without a second pass.
    columns : dict of str to int
        Field name to worksheet column index, for the fields this year publishes.
    unknown_columns : list of str
        Headers that are not in :data:`COLUMNS`, as published. Reported by the
        importer rather than raised on: a new column is a reason to look, not a
        reason to refuse a year of data.
    records : iterator of FireRecord
        The rows, streamed. Consumable once.
    """

    year: int
    name: str
    header_row: int
    rows: int
    columns: dict[str, int]
    unknown_columns: list[str]
    records: typing.Iterator[FireRecord]


def clean(value: object) -> str | None:
    """Strip a published cell to text, or ``None`` if it holds nothing.

    The Greek workbooks use ordinary empty cells rather than the non-breaking
    spaces the EGIF exports write, but they do leave whitespace-only ones behind —
    ``Δασαρχείο`` is blank on most rows and is sometimes a run of spaces.
    """
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_number(value: object) -> float | None:
    """A published area or count as a number, or ``None``.

    Cells arrive as ``int`` or ``float`` from ``openpyxl`` — 255,721 and 4,472 of
    them respectively over the archive — but text is accepted too, because a
    single re-typed cell in a future publication is not a reason to lose a year.
    A comma decimal separator is read as one: Greek locale writes ``0,9``.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_integer(value: object) -> int | None:
    """A published identifier or count as an ``int``, or ``None``.

    ``0`` is preserved for a count — no aircraft flew to this fire is an answer —
    but see :func:`to_identifier` for why it is not preserved for an identifier.
    """
    number = to_number(value)
    return None if number is None else int(number)


def to_identifier(value: object) -> int | None:
    """A published ``Α/Α ΕΓΓΡΑΦΗΣ`` or ``Α/Α ENGAGE``, or ``None``.

    ``0`` becomes ``None``. The service writes it in ``Α/Α ENGAGE`` for a record
    its dispatch system has no incident for — the same rows whose coordinates are
    ``0``/``0`` — and storing a literal zero would make 3,755 unrelated fires share
    an identifier.
    """
    identifier = to_integer(value)
    return None if not identifier else identifier


def to_date(value: object) -> datetime.date | None:
    """A published date cell, whichever of the two forms it is in.

    A real Excel datetime for 2000-2024, ``dd/mm/yyyy`` text for 2025. Anything
    else returns ``None`` and the caller records the problem.
    """
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    text = clean(value)
    if text is None:
        return None
    match = TEXT_DATE.match(text)
    if match is None:
        return None
    day, month, year = (int(group) for group in match.groups())
    try:
        return datetime.date(year, month, day)
    except ValueError:
        return None


def to_time(value: object) -> datetime.time | None:
    """A published time cell, whichever of the two forms it is in.

    A real Excel time for 2000-2006 and 2008-2010, ``HH:MM`` text everywhere else.

    A published ``24:00`` is read as ``00:00``, which is what the service means by
    it — midnight at the end of the day — and is why the hour is checked here
    rather than handed to :class:`datetime.time`, which would refuse it.
    """
    if isinstance(value, datetime.datetime):
        return value.time()
    if isinstance(value, datetime.time):
        return value
    text = clean(value)
    if text is None:
        return None
    match = TEXT_TIME.match(text)
    if match is None:
        return None
    hour, minute, second = (int(group or 0) for group in match.groups())
    if hour == 24 and minute == 0:
        hour = 0
    try:
        return datetime.time(hour, minute, second)
    except ValueError:
        return None


def find_header(rows: list[tuple[object, ...]]) -> int | None:
    """Index into ``rows`` of the column-name row, or ``None``.

    The first row naming at least :data:`MIN_KNOWN_COLUMNS` columns this reader
    knows. A property of the row and not of the file, which is what lets one rule
    serve a header on row 1 (2014), row 2 (twenty-four sheets) and row 4 (2025).
    """
    for index, row in enumerate(rows):
        known = sum(1 for cell in row if normalise_column(str(cell) if cell else "") in COLUMNS)
        if known >= MIN_KNOWN_COLUMNS:
            return index
    return None


def find_year(name: str, rows: list[tuple[object, ...]]) -> int | None:
    """The year a sheet is filed under, or ``None`` if it cannot be established.

    The sheet name when it is four digits, which covers every published file but
    one. Otherwise the first plausible year in the cells above the header, which is
    where ``agrotodasikes_pyrkaies_2025.xlsx`` puts it — ``Για το ΕΤΟΣ:`` and then
    ``2025`` in the cell beside it.

    Parameters
    ----------
    name : str
        The worksheet name.
    rows : list of tuple
        The rows above the header, to search when the name does not say.
    """
    if name.strip().isdigit() and len(name.strip()) == 4:
        year = int(name.strip())
        if PLAUSIBLE_YEARS[0] <= year <= PLAUSIBLE_YEARS[1]:
            return year
    for row in rows:
        for cell in row:
            year = to_integer(cell) if not isinstance(cell, str) else None
            if year is None and isinstance(cell, str) and cell.strip().isdigit():
                year = int(cell.strip())
            if year is not None and PLAUSIBLE_YEARS[0] <= year <= PLAUSIBLE_YEARS[1]:
                return year
    return None


def read_row(values: tuple[object, ...], columns: dict[str, int], row: int) -> FireRecord:
    """Convert one worksheet row into a :class:`FireRecord`.

    Each field is converted by what it is, not by where it sits: identifiers and
    counts through :func:`to_identifier` and :func:`to_integer`, areas through
    :func:`to_number`, dates and times through :func:`to_date` and :func:`to_time`,
    and everything else as text.
    """
    def cell(field: str) -> object:
        index = columns.get(field)
        if index is None or index >= len(values):
            return None
        return values[index]

    record = FireRecord(row=row)

    record.record_number = to_identifier(cell("record_number"))
    record.engage_id = to_identifier(cell("engage_id"))
    record.incident_category = clean(cell("incident_category"))

    for field in ("start_date", "end_date"):
        raw = cell(field)
        parsed = to_date(raw)
        if parsed is None and clean(raw) is not None:
            record.problems.append(f"{field} is not a date: {clean(raw)!r}")
        setattr(record, field, parsed)
    for field in ("start_time", "end_time"):
        raw = cell(field)
        parsed = to_time(raw)
        if parsed is None and clean(raw) is not None:
            record.problems.append(f"{field} is not a time: {clean(raw)!r}")
        setattr(record, field, parsed)

    record.longitude = to_number(cell("longitude"))
    record.latitude = to_number(cell("latitude"))
    for field in ("station_name", "prefecture_name", "forest_district_name",
                  "municipality_name", "locality_name", "address"):
        setattr(record, field, clean(cell(field)))

    for field in AREA_FIELDS:
        setattr(record, field, to_number(cell(field)))
    for field in COUNT_FIELDS:
        setattr(record, field, to_integer(cell(field)))

    return record


def _records(worksheet, columns: dict[str, int],
             header_row: int) -> typing.Iterator[FireRecord]:
    """Stream the data rows of one worksheet, skipping the wholly empty ones."""
    for number, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1):
        if all(value is None for value in values):
            continue
        yield read_row(values, columns, number)


def read_workbook(path: Path) -> typing.Iterator[Sheet]:
    """Yield one :class:`Sheet` per year of fires in a workbook.

    Sheets in :data:`IGNORED_SHEETS` are skipped, as is any sheet with no
    recognisable header — a publication note, a pivot table, an empty tab.

    Parameters
    ----------
    path : pathlib.Path
        The ``.xlsx`` to read.

    Yields
    ------
    Sheet
        One per year found, in workbook order. Each sheet's ``records`` must be
        consumed before the next sheet is taken, since both stream from the same
        open workbook.

    Raises
    ------
    RuntimeError
        If the workbook cannot be opened at all, or if a sheet has a header this
        reader recognises but no year — the year is what an import replaces, and
        replacing the wrong one would delete a year of good data.
    """
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001  (openpyxl raises a zoo of exceptions)
        raise RuntimeError(f"{path.name} is not a readable .xlsx: {error}") from error

    try:
        for worksheet in workbook.worksheets:
            if normalise_column(worksheet.title) in IGNORED_SHEETS:
                continue

            preamble = [values for _, values in zip(
                range(MAX_HEADER_ROW),
                worksheet.iter_rows(min_row=1, max_row=MAX_HEADER_ROW, values_only=True),
            )]
            offset = find_header(preamble)
            if offset is None:
                continue
            header_row = offset + 1

            year = find_year(worksheet.title, preamble[:offset])
            if year is None:
                raise RuntimeError(
                    f"{path.name}: sheet {worksheet.title!r} has a header but no year, "
                    f"neither in its name nor above the header. The year is what an "
                    f"import replaces, so it cannot be guessed."
                )

            columns: dict[str, int] = {}
            unknown: list[str] = []
            for index, cell in enumerate(preamble[offset]):
                name = clean(cell)
                if name is None:
                    continue
                field = COLUMNS.get(normalise_column(name))
                if field is None:
                    unknown.append(name)
                elif field not in columns:
                    columns[field] = index

            yield Sheet(
                year=year,
                name=worksheet.title,
                header_row=header_row,
                rows=max((worksheet.max_row or header_row) - header_row, 0),
                columns=columns,
                unknown_columns=unknown,
                records=_records(worksheet, columns, header_row),
            )
    finally:
        workbook.close()
